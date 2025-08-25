from django.db.models import Sum, F, Count, DecimalField, ExpressionWrapper
from django.utils.timezone import now
from datetime import timedelta
from django.shortcuts import render, redirect
from django.contrib import messages
from django.http import HttpResponse
from inventario.models import MovimientoInventario, Inventario
from productos.models import Producto, Clase
from categorias.models import Categoria, Subcategoria
from usuarios.models import Sucursal
from usuarios.views import obtener_empresa
from usuarios.templatetags.tags import control_acceso
from decimal import Decimal
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO


@control_acceso('Auditor')
def reporte_movimientos_dia(request):
    # Obtener empresa del usuario
    empresa_actual = obtener_empresa(request)
    
    # Obtener la fecha de hoy
    fecha = now() - timedelta(days=0)
    inicio_dia = fecha.replace(hour=0, minute=0, second=0, microsecond=0)
    fin_dia = fecha.replace(hour=23, minute=59, second=59, microsecond=999999)

    # Filtrar los movimientos del día actual por empresa
    movimientos = MovimientoInventario.objects.filter(
        fecha__gte=inicio_dia, 
        fecha__lte=fin_dia,
        producto__empresa=empresa_actual
    )

    # Agrupar movimientos por usuario y resumir cantidades
    resumen_por_usuario = movimientos.values('usuario__username').annotate(
        total_cantidad=Sum('cantidad'),
        total_movimientos=Count('id')
    )

    # Agrupar movimientos por tipo y resumir cantidades
    resumen_por_tipo = movimientos.values('tipo_movimiento').annotate(
        total_cantidad=Sum('cantidad'),
        total_movimientos=Count('id')
    )

    # Detalles completos para la tabla
    detalles_movimientos = movimientos.select_related('producto', 'sucursal').order_by('-fecha')

    return render(request, 'reporte_movimientos_dia.html', {
        'fecha_hoy': fecha,
        'resumen_por_usuario': resumen_por_usuario,
        'resumen_por_tipo': resumen_por_tipo,
        'detalles_movimientos': detalles_movimientos,
        'empresa': empresa_actual,
    })


@control_acceso('Auditor')
def reporte_inventario_valorizado(request):
    """
    Reporte de inventario valorizado mostrando el valor total del inventario
    por producto, sucursal y categoría.
    """
    empresa_actual = obtener_empresa(request)
    
    # Obtener inventarios con stock > 0
    inventarios = Inventario.objects.filter(
        producto__empresa=empresa_actual,
        cantidad__gt=0
    ).select_related(
        'producto', 
        'producto__clase__subcategoria__categoria',
        'sucursal'
    ).annotate(
        valor_total=ExpressionWrapper(
            F('cantidad') * F('producto__precio'),
            output_field=DecimalField(max_digits=10, decimal_places=2)
        )
    ).order_by('sucursal__nombre', 'producto__nombre')
    
    # Calcular totales por sucursal
    totales_por_sucursal = inventarios.values(
        'sucursal__id', 
        'sucursal__nombre'
    ).annotate(
        total_productos=Count('producto', distinct=True),
        total_unidades=Sum('cantidad'),
        valor_total=Sum(
            ExpressionWrapper(
                F('cantidad') * F('producto__precio'),
                output_field=DecimalField(max_digits=10, decimal_places=2)
            )
        )
    ).order_by('sucursal__nombre')
    
    # Calcular totales por categoría
    totales_por_categoria = inventarios.values(
        'producto__clase__subcategoria__categoria__id',
        'producto__clase__subcategoria__categoria__nombre'
    ).annotate(
        total_productos=Count('producto', distinct=True),
        total_unidades=Sum('cantidad'),
        valor_total=Sum(
            ExpressionWrapper(
                F('cantidad') * F('producto__precio'),
                output_field=DecimalField(max_digits=10, decimal_places=2)
            )
        )
    ).order_by('-valor_total')
    
    # Productos con mayor valor en inventario
    productos_top = inventarios.values(
        'producto__id',
        'producto__codigo',
        'producto__nombre'
    ).annotate(
        stock_total=Sum('cantidad'),
        precio=F('producto__precio'),
        valor_total=Sum(
            ExpressionWrapper(
                F('cantidad') * F('producto__precio'),
                output_field=DecimalField(max_digits=10, decimal_places=2)
            )
        )
    ).order_by('-valor_total')[:10]
    
    # Productos con stock bajo
    productos_stock_bajo = Inventario.objects.filter(
        producto__empresa=empresa_actual,
        cantidad__lte=F('stock_minimo')
    ).select_related('producto', 'sucursal').order_by('cantidad')
    
    # Calcular totales generales
    total_general = inventarios.aggregate(
        total_productos=Count('producto', distinct=True),
        total_unidades=Sum('cantidad'),
        valor_total=Sum(
            ExpressionWrapper(
                F('cantidad') * F('producto__precio'),
                output_field=DecimalField(max_digits=10, decimal_places=2)
            )
        )
    )
    
    return render(request, 'reporte_inventario_valorizado.html', {
        'inventarios': inventarios,
        'totales_por_sucursal': totales_por_sucursal,
        'totales_por_categoria': totales_por_categoria,
        'productos_top': productos_top,
        'productos_stock_bajo': productos_stock_bajo,
        'total_general': total_general,
        'empresa': empresa_actual,
        'fecha_reporte': now()
    })


@control_acceso('Manaudi')
def descargar_plantilla_productos(request):
    """
    Descarga la plantilla Excel para carga masiva de productos con datos reales de la empresa
    """
    empresa_actual = obtener_empresa(request)
    
    # Crear un libro de Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Plantilla Productos"
    
    # Definir encabezados y sus descripciones
    headers = [
        ('codigo', 'Código único del producto (obligatorio)'),
        ('nombre', 'Nombre del producto (obligatorio)'),
        ('modelo', 'Modelo del producto (opcional)'),
        ('categoria', 'Nombre de la categoría (obligatorio)'),
        ('subcategoria', 'Nombre de la subcategoría (obligatorio)'),
        ('clase', 'Nombre de la clase (obligatorio)'),
        ('precio', 'Precio del producto (número decimal)'),
        ('tipo_producto', 'Tipo: unidad o juego'),
        ('descripcion', 'Descripción del producto (opcional)'),
        ('stock_minimo', 'Stock mínimo para alertas (número entero)'),
        ('sucursal', 'Nombre de la sucursal para inventario inicial'),
        ('cantidad_inicial', 'Cantidad inicial en inventario (número entero)')
    ]
    
    # Estilos
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Escribir encabezados
    for col, (header, description) in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Añadir comentario con descripción
        from openpyxl.comments import Comment
        cell.comment = Comment(description, "Sistema")
        
        # Ajustar ancho de columna
        ws.column_dimensions[get_column_letter(col)].width = 15
    
    # Obtener datos reales de la empresa
    categorias = Categoria.objects.filter(empresa=empresa_actual)
    sucursales = Sucursal.objects.filter(empresa=empresa_actual)
    
    # Generar ejemplos con datos reales
    ejemplos = []
    row_count = 2
    
    # Intentar crear ejemplos con datos reales
    for categoria in categorias[:3]:  # Limitar a 3 categorías para los ejemplos
        subcategorias = Subcategoria.objects.filter(categoria=categoria)[:2]  # 2 subcategorías por categoría
        for subcategoria in subcategorias:
            clases = Clase.objects.filter(subcategoria=subcategoria)[:2]  # 2 clases por subcategoría
            for clase in clases:
                if row_count <= 10:  # Limitar a 10 filas de ejemplo
                    sucursal_nombre = sucursales.first().nombre if sucursales else 'Sucursal Principal'
                    ejemplo = [
                        f'PROD{str(row_count-1).zfill(3)}',  # Código de ejemplo
                        f'Producto de ejemplo {row_count-1}',  # Nombre de ejemplo
                        f'MOD-{str(row_count-1).zfill(3)}',  # Modelo de ejemplo
                        categoria.nombre,
                        subcategoria.nombre,
                        clase.nombre,
                        '0.00',  # Precio
                        'unidad',  # Tipo producto
                        f'Descripción del producto {row_count-1}',  # Descripción
                        '10',  # Stock mínimo
                        sucursal_nombre,  # Sucursal
                        '0'  # Cantidad inicial
                    ]
                    ejemplos.append(ejemplo)
                    row_count += 1
    
    # Si no hay datos reales, usar ejemplos genéricos
    if not ejemplos:
        sucursal_nombre = sucursales.first().nombre if sucursales else 'Sucursal Principal'
        ejemplos = [
            ['PROD001', 'Producto de ejemplo 1', 'MOD-001', 'CATEGORIA1', 'SUBCATEGORIA1', 'CLASE1', '45.50', 'unidad', 'Descripción del producto 1', '10', sucursal_nombre, '50'],
            ['PROD002', 'Producto de ejemplo 2', 'MOD-002', 'CATEGORIA1', 'SUBCATEGORIA1', 'CLASE2', '12.75', 'unidad', 'Descripción del producto 2', '20', sucursal_nombre, '100'],
            ['PROD003', 'Producto de ejemplo 3', 'MOD-003', 'CATEGORIA2', 'SUBCATEGORIA2', 'CLASE1', '35.00', 'juego', 'Descripción del producto 3', '5', sucursal_nombre, '25'],
        ]
    
    # Escribir ejemplos en la plantilla
    for row_idx, ejemplo in enumerate(ejemplos, 2):
        for col_idx, valor in enumerate(ejemplo, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=valor)
            cell.border = border
            if row_idx % 2 == 0:
                cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    
    # Crear hoja de instrucciones
    ws2 = wb.create_sheet("Instrucciones")
    instrucciones = [
        ["INSTRUCCIONES PARA CARGA MASIVA DE PRODUCTOS"],
        [""],
        ["1. CAMPOS OBLIGATORIOS:"],
        ["   - codigo: Debe ser único para cada producto en la empresa"],
        ["   - nombre: Nombre descriptivo del producto"],
        ["   - categoria: Use valores de la hoja 'Categorías' o se creará automáticamente"],
        ["   - subcategoria: Use valores de la hoja 'Subcategorías' o se creará automáticamente"],
        ["   - clase: Use valores de la hoja 'Clases' o se creará automáticamente"],
        [""],
        ["2. CAMPOS OPCIONALES:"],
        ["   - modelo: Modelo del producto (se guardará en mayúsculas)"],
        ["   - precio: Por defecto será 0.00 si no se especifica"],
        ["   - tipo_producto: 'unidad' o 'juego'. Por defecto será 'unidad'"],
        ["   - descripcion: Texto libre para describir el producto"],
        ["   - stock_minimo: Por defecto será 0 si no se especifica"],
        ["   - sucursal: Use valores de la hoja 'Sucursales'"],
        ["   - cantidad_inicial: Requiere que se especifique la sucursal"],
        [""],
        ["3. NOTAS IMPORTANTES:"],
        ["   - No modifique los nombres de las columnas"],
        ["   - Los códigos de producto deben ser únicos"],
        ["   - Revise las hojas de catálogos para usar valores existentes"],
        ["   - Los precios deben ser números decimales (use punto como separador)"],
        ["   - Las cantidades deben ser números enteros"],
        ["   - Elimine las filas de ejemplo antes de cargar sus datos"],
        [""],
        ["4. VALORES VÁLIDOS:"],
        ["   - tipo_producto: 'unidad' o 'juego'"],
        ["   - precio: Números positivos con hasta 2 decimales"],
        ["   - cantidad_inicial y stock_minimo: Números enteros positivos"],
        [""],
        ["5. HOJAS DE REFERENCIA:"],
        ["   - Categorías: Lista de categorías disponibles"],
        ["   - Subcategorías: Lista de subcategorías por categoría"],
        ["   - Clases: Lista de clases por subcategoría"],
        ["   - Sucursales: Lista de sucursales disponibles"],
    ]
    
    for row_idx, instruccion in enumerate(instrucciones, 1):
        if instruccion:
            cell = ws2.cell(row=row_idx, column=1, value=instruccion[0])
            if row_idx == 1:
                cell.font = Font(bold=True, size=14)
            elif instruccion[0].startswith(("1.", "2.", "3.", "4.", "5.")):
                cell.font = Font(bold=True, size=11)
    
    # Ajustar ancho de columna de instrucciones
    ws2.column_dimensions['A'].width = 80
    
    # Crear hoja de Categorías
    ws3 = wb.create_sheet("Categorías")
    ws3.cell(row=1, column=1, value="CATEGORÍA").font = Font(bold=True)
    ws3.cell(row=1, column=1).fill = PatternFill(start_color="E6E6E6", end_color="E6E6E6", fill_type="solid")
    for idx, categoria in enumerate(categorias, 2):
        ws3.cell(row=idx, column=1, value=categoria.nombre)
    ws3.column_dimensions['A'].width = 30
    
    # Crear hoja de Subcategorías
    ws4 = wb.create_sheet("Subcategorías")
    ws4.cell(row=1, column=1, value="CATEGORÍA").font = Font(bold=True)
    ws4.cell(row=1, column=2, value="SUBCATEGORÍA").font = Font(bold=True)
    ws4.cell(row=1, column=1).fill = PatternFill(start_color="E6E6E6", end_color="E6E6E6", fill_type="solid")
    ws4.cell(row=1, column=2).fill = PatternFill(start_color="E6E6E6", end_color="E6E6E6", fill_type="solid")
    
    row_idx = 2
    for categoria in categorias:
        subcategorias = Subcategoria.objects.filter(categoria=categoria)
        for subcategoria in subcategorias:
            ws4.cell(row=row_idx, column=1, value=categoria.nombre)
            ws4.cell(row=row_idx, column=2, value=subcategoria.nombre)
            row_idx += 1
    ws4.column_dimensions['A'].width = 30
    ws4.column_dimensions['B'].width = 30
    
    # Crear hoja de Clases
    ws5 = wb.create_sheet("Clases")
    ws5.cell(row=1, column=1, value="CATEGORÍA").font = Font(bold=True)
    ws5.cell(row=1, column=2, value="SUBCATEGORÍA").font = Font(bold=True)
    ws5.cell(row=1, column=3, value="CLASE").font = Font(bold=True)
    for col in range(1, 4):
        ws5.cell(row=1, column=col).fill = PatternFill(start_color="E6E6E6", end_color="E6E6E6", fill_type="solid")
    
    row_idx = 2
    for categoria in categorias:
        subcategorias = Subcategoria.objects.filter(categoria=categoria)
        for subcategoria in subcategorias:
            clases = Clase.objects.filter(subcategoria=subcategoria)
            for clase in clases:
                ws5.cell(row=row_idx, column=1, value=categoria.nombre)
                ws5.cell(row=row_idx, column=2, value=subcategoria.nombre)
                ws5.cell(row=row_idx, column=3, value=clase.nombre)
                row_idx += 1
    ws5.column_dimensions['A'].width = 30
    ws5.column_dimensions['B'].width = 30
    ws5.column_dimensions['C'].width = 30
    
    # Crear hoja de Sucursales
    ws6 = wb.create_sheet("Sucursales")
    ws6.cell(row=1, column=1, value="SUCURSAL").font = Font(bold=True)
    ws6.cell(row=1, column=2, value="TIPO").font = Font(bold=True)
    ws6.cell(row=1, column=1).fill = PatternFill(start_color="E6E6E6", end_color="E6E6E6", fill_type="solid")
    ws6.cell(row=1, column=2).fill = PatternFill(start_color="E6E6E6", end_color="E6E6E6", fill_type="solid")
    
    for idx, sucursal in enumerate(sucursales, 2):
        ws6.cell(row=idx, column=1, value=sucursal.nombre)
        ws6.cell(row=idx, column=2, value=sucursal.get_tipo_sucursal_display())
    ws6.column_dimensions['A'].width = 30
    ws6.column_dimensions['B'].width = 20
    
    # Guardar en memoria
    virtual_workbook = BytesIO()
    wb.save(virtual_workbook)
    virtual_workbook.seek(0)
    
    # Preparar respuesta
    response = HttpResponse(
        virtual_workbook.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=plantilla_productos.xlsx'
    
    return response


@control_acceso('Manaudi')
def carga_masiva_productos(request):
    """
    Vista para cargar productos masivamente desde un archivo Excel
    """
    empresa_actual = obtener_empresa(request)
    
    if request.method == 'POST' and request.FILES.get('archivo_excel'):
        archivo = request.FILES['archivo_excel']
        
        # Validar extensión
        if not archivo.name.endswith(('.xlsx', '.xls')):
            messages.error(request, 'Por favor suba un archivo Excel válido (.xlsx o .xls)')
            return redirect('carga_masiva_productos')
        
        try:
            # Leer el archivo Excel
            wb = openpyxl.load_workbook(archivo)
            ws = wb.active
            
            # Obtener encabezados
            headers = []
            for cell in ws[1]:
                headers.append(cell.value)
            
            # Validar encabezados requeridos
            required_headers = ['codigo', 'nombre', 'categoria', 'subcategoria', 'clase']
            for header in required_headers:
                if header not in headers:
                    messages.error(request, f'Falta la columna requerida: {header}')
                    return redirect('carga_masiva_productos')
            
            # Procesar filas
            productos_creados = 0
            productos_actualizados = 0
            errores = []
            
            for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                # Crear diccionario con los datos
                data = {}
                for idx, header in enumerate(headers):
                    if header and idx < len(row):
                        data[header] = row[idx]
                
                # Saltar filas vacías
                if not data.get('codigo') or not data.get('nombre'):
                    continue
                
                try:
                    # Obtener o crear categoría
                    categoria, _ = Categoria.objects.get_or_create(
                        nombre=str(data['categoria']).upper(),
                        empresa=empresa_actual
                    )
                    
                    # Obtener o crear subcategoría
                    subcategoria, _ = Subcategoria.objects.get_or_create(
                        nombre=str(data['subcategoria']).upper(),
                        categoria=categoria
                    )
                    
                    # Obtener o crear clase
                    clase, _ = Clase.objects.get_or_create(
                        nombre=str(data['clase']).upper(),
                        subcategoria=subcategoria
                    )
                    
                    # Preparar datos del producto
                    producto_data = {
                        'nombre': data['nombre'],
                        'modelo': data.get('modelo', ''),
                        'clase': clase,
                        'empresa': empresa_actual,
                        'precio': Decimal(str(data.get('precio', 0))),
                        'tipo_producto': data.get('tipo_producto', 'unidad').lower(),
                        'descripcion': data.get('descripcion', ''),
                        'estado': 'activo'
                    }
                    
                    # Validar tipo_producto
                    if producto_data['tipo_producto'] not in ['unidad', 'juego']:
                        producto_data['tipo_producto'] = 'unidad'
                    
                    # Crear o actualizar producto
                    producto, created = Producto.objects.update_or_create(
                        codigo=data['codigo'],
                        empresa=empresa_actual,
                        defaults=producto_data
                    )
                    
                    if created:
                        productos_creados += 1
                    else:
                        productos_actualizados += 1
                    
                    # Crear inventario inicial si se especifica
                    if data.get('sucursal') and data.get('cantidad_inicial'):
                        try:
                            sucursal = Sucursal.objects.get(
                                nombre__iexact=data['sucursal'],
                                empresa=empresa_actual
                            )
                            
                            inventario, _ = Inventario.objects.update_or_create(
                                producto=producto,
                                sucursal=sucursal,
                                defaults={
                                    'cantidad': int(data['cantidad_inicial']),
                                    'stock_minimo': int(data.get('stock_minimo', 0))
                                }
                            )
                        except Sucursal.DoesNotExist:
                            errores.append(f"Fila {row_num}: Sucursal '{data['sucursal']}' no encontrada")
                        except ValueError:
                            errores.append(f"Fila {row_num}: Cantidad inicial debe ser un número entero")
                    
                except Exception as e:
                    errores.append(f"Fila {row_num}: {str(e)}")
            
            # Mostrar resumen
            if productos_creados > 0:
                messages.success(request, f'Se crearon {productos_creados} productos nuevos')
            if productos_actualizados > 0:
                messages.info(request, f'Se actualizaron {productos_actualizados} productos existentes')
            if errores:
                for error in errores[:10]:  # Mostrar máximo 10 errores
                    messages.warning(request, error)
                if len(errores) > 10:
                    messages.warning(request, f'... y {len(errores) - 10} errores más')
            
            if productos_creados == 0 and productos_actualizados == 0 and not errores:
                messages.warning(request, 'No se encontraron datos válidos para procesar')
            
        except Exception as e:
            messages.error(request, f'Error al procesar el archivo: {str(e)}')
        
        return redirect('carga_masiva_productos')
    
    # Obtener resumen de datos actuales
    resumen = {
        'total_productos': Producto.objects.filter(empresa=empresa_actual).count(),
        'total_categorias': Categoria.objects.filter(empresa=empresa_actual).count(),
        'total_subcategorias': Subcategoria.objects.filter(categoria__empresa=empresa_actual).count(),
        'total_clases': Clase.objects.filter(subcategoria__categoria__empresa=empresa_actual).count(),
        'sucursales': Sucursal.objects.filter(empresa=empresa_actual).values_list('nombre', flat=True)
    }
    
    return render(request, 'carga_masiva_productos.html', {
        'empresa': empresa_actual,
        'resumen': resumen
    })